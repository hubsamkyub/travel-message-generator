import streamlit as st
import pandas as pd
import re
from datetime import datetime

def show_success_metric(title, value, delta=None):
    """성공 메트릭 표시"""
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.metric(
            label=title,
            value=value,
            delta=delta
        )

def show_processing_steps(steps, current_step=0):
    """처리 단계 시각화"""
    cols = st.columns(len(steps))
    
    for i, (col, step) in enumerate(zip(cols, steps)):
        with col:
            if i < current_step:
                st.success(f"✅ {step}")
            elif i == current_step:
                st.info(f"🔄 {step}")
            else:
                st.write(f"⏳ {step}")

def create_info_card(title, content, icon="ℹ️", type="info"):
    """정보 카드 생성"""
    if type == "success":
        st.success(f"{icon} **{title}**\n\n{content}")
    elif type == "warning":
        st.warning(f"{icon} **{title}**\n\n{content}")
    elif type == "error":
        st.error(f"{icon} **{title}**\n\n{content}")
    else:
        st.info(f"{icon} **{title}**\n\n{content}")

def show_data_summary(df, title="데이터 요약"):
    """데이터 요약 정보 표시"""
    with st.expander(f"📊 {title}", expanded=False):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("총 행 수", len(df))
        with col2:
            st.metric("총 열 수", len(df.columns))
        with col3:
            null_count = df.isnull().sum().sum()
            st.metric("빈 셀 수", null_count)
        
        st.markdown("**컬럼 정보:**")
        for i, col in enumerate(df.columns):
            non_null = df[col].notna().sum()
            st.write(f"• {col}: {non_null}/{len(df)} 값")

def highlight_template_variables(template_text):
    """템플릿 변수 하이라이팅 HTML 생성"""
    # 변수 패턴 찾기 {변수명} 또는 {변수명:포맷}
    pattern = r'\{([^}]+)\}'
    
    def replace_var(match):
        var_content = match.group(1)
        var_name = var_content.split(':')[0]  # 포맷 부분 제거
        return f'<span style="background-color: #fff3cd; padding: 2px 4px; border-radius: 3px; font-family: monospace; color: #856404;">{{{var_content}}}</span>'
    
    highlighted = re.sub(pattern, replace_var, template_text)
    return highlighted

# ui_helpers.py 파일에서 show_template_preview 함수를 아래 코드로 교체하세요.

def show_template_preview(template, preview_data):
    """
    템플릿 미리보기 (실제 데이터를 받아 미리보기를 생성)
    """
    st.markdown("##### 🔍 템플릿 미리보기")

    # preview_data가 없는 경우를 대비한 안전장치
    if not preview_data:
        st.warning("미리보기를 생성할 데이터가 없습니다. 매핑 설정을 확인해주세요.")
        # 빈 박스라도 보여주어 레이아웃 유지
        st.text_area("Template Preview Area", "미리보기 생성 불가", height=500, disabled=True, label_visibility="collapsed")
        return

    try:
        # 템플릿에만 있고 데이터에는 없는 변수들을 위해 기본값 추가
        template_vars = set(re.findall(r'\{(\w+)(?::[^}]+)?\}', template))
        for var in template_vars:
            if var not in preview_data:
                preview_data[var] = f"[{var}]"

        # 숫자 포맷팅({var:,})이 사용된 변수는 숫자 타입으로 변환 (오류 방지)
        number_format_vars = set(re.findall(r'\{(\w+):[^}]*[,d][^}]*\}', template))
        for var_name in number_format_vars:
            if var_name in preview_data:
                try:
                    # 값이 문자열일 경우, 숫자 부분만 추출하여 변환
                    if isinstance(preview_data[var_name], str):
                        cleaned_val = re.sub(r'[^\d.-]', '', preview_data[var_name])
                        preview_data[var_name] = int(float(cleaned_val)) if cleaned_val else 0
                    elif not isinstance(preview_data[var_name], (int, float)):
                        preview_data[var_name] = 0
                except (ValueError, TypeError):
                    preview_data[var_name] = 0 # 변환 실패 시 0으로 처리

        preview_message = template.format(**preview_data)

        # st.text_area를 사용하여 좌우 대칭 UI 및 안정적인 미리보기 제공
        st.text_area(
            label="Template Preview Area",
            value=preview_message,
            height=500,
            disabled=True,
            label_visibility="collapsed",
            help="실제 엑셀의 첫 번째 데이터를 기반으로 생성된 미리보기입니다."
        )

    except KeyError as e:
        missing_var = str(e).strip("'")
        st.error(f"❌ 템플릿 오류: 정의되지 않은 변수 {{{missing_var}}}가 사용되었습니다.")
    except Exception as e:
        st.error(f"❌ 미리보기 생성 중 오류가 발생했습니다: {str(e)}")
        
def show_variable_suggestions(df_columns):
    """변수 제안 표시"""
    st.markdown("**💡 사용 가능한 변수 제안:**")
    
    # 기본 변수들
    basic_vars = [
        "product_name", "payment_due_date", "base_exchange_rate", 
        "current_exchange_rate", "exchange_rate_diff", "company_burden"
    ]
    
    # 그룹 변수들
    group_vars = [
        "team_name", "sender_group", "group_members_text", 
        "group_size", "total_balance", "bank_account"
    ]
    
    # 컬럼에서 추정되는 변수들
    suggested_vars = []
    for col in df_columns:
        col_lower = str(col).lower()
        if any(keyword in col_lower for keyword in ['금액', '요금', '비용', 'price', 'fee', 'amount']):
            var_name = col.replace(' ', '_').replace('-', '_').lower()
            suggested_vars.append(f"{var_name} (from '{col}')")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("**기본 정보:**")
        for var in basic_vars:
            st.code(f"{{{var}}}")
    
    with col2:
        st.markdown("**그룹 정보:**")
        for var in group_vars:
            st.code(f"{{{var}}}")
    
    with col3:
        st.markdown("**엑셀 컬럼 기반:**")
        for var in suggested_vars[:5]:  # 최대 5개만 표시
            st.code(f"{{{var.split(' ')[0]}}}")

def show_group_statistics(group_data):
    """그룹 통계 표시"""
    if not group_data:
        return
    
    st.markdown("**📈 그룹 통계**")
    
    # 기본 통계
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("총 그룹 수", len(group_data))
    
    with col2:
        total_members = sum(group['group_size'] for group in group_data.values())
        st.metric("총 인원 수", total_members)
    
    with col3:
        teams = set(group['team_name'] for group in group_data.values())
        st.metric("팀 수", len(teams))
    
    with col4:
        avg_size = total_members / len(group_data) if group_data else 0
        st.metric("평균 그룹 크기", f"{avg_size:.1f}명")
    
    # 팀별 분포
    team_distribution = {}
    for group in group_data.values():
        team = group['team_name']
        team_distribution[team] = team_distribution.get(team, 0) + 1
    
    if team_distribution:
        st.markdown("**팀별 그룹 분포:**")
        chart_data = pd.DataFrame(
            list(team_distribution.items()),
            columns=['팀', '그룹 수']
        )
        st.bar_chart(chart_data.set_index('팀'))

def create_download_section(generated_messages):
    """다운로드 섹션 생성"""
    st.markdown("---")
    st.markdown("## 📥 결과 다운로드")
    
    if not generated_messages:
        st.warning("생성된 메시지가 없습니다.")
        return
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # 텍스트 파일 다운로드
        txt_content = create_text_download_content(generated_messages)
        st.download_button(
            label="📄 텍스트 파일 다운로드",
            data=txt_content,
            file_name=f"messages_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            mime="text/plain",
            help="모든 메시지를 텍스트 파일로 다운로드합니다."
        )
    
    with col2:
        # 엑셀 파일 다운로드
        excel_content = create_excel_download_content(generated_messages)
        st.download_button(
            label="📊 엑셀 파일 다운로드", 
            data=excel_content,
            file_name=f"messages_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="그룹 정보와 메시지를 엑셀 파일로 다운로드합니다."
        )
    
    with col3:
        # CSV 파일 다운로드
        csv_content = create_csv_download_content(generated_messages)
        st.download_button(
            label="📋 CSV 파일 다운로드",
            data=csv_content,
            file_name=f"messages_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            help="그룹 정보를 CSV 파일로 다운로드합니다."
        )

def create_text_download_content(generated_messages):
    """텍스트 다운로드 컨텐츠 생성"""
    content = []
    content.append(f"여행 잔금 문자 메시지")
    content.append(f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    content.append(f"총 {len(generated_messages)}개 그룹")
    content.append("=" * 60)
    content.append("")
    
    for group_id, data in generated_messages.items():
        group_info = data['group_info']
        message = data['message']
        
        content.append(f"[{group_id}] {group_info['team_name']} - {group_info['sender_group']}")
        content.append(f"발송인: {group_info['sender']}")
        content.append(f"대상자: {', '.join(group_info['members'])} ({group_info['group_size']}명)")
        content.append(f"연락처: {group_info.get('contact', '')}")
        content.append("-" * 40)
        content.append(message)
        content.append("")
        content.append("=" * 60)
        content.append("")
    
    return "\n".join(content)

def create_excel_download_content(generated_messages):
    """엑셀 다운로드 컨텐츠 생성"""
    import io
    
    data = []
    for group_id, message_data in generated_messages.items():
        group_info = message_data['group_info']
        message = message_data['message']
        
        data.append({
            '그룹ID': group_id,
            '팀명': group_info['team_name'],
            '발송그룹': group_info['sender_group'],
            '발송인': group_info['sender'],
            '연락처': group_info.get('contact', ''),
            '그룹멤버': ', '.join(group_info['members']),
            '인원수': group_info['group_size'],
            '총잔금': group_info.get('total_balance', ''),
            '메시지': message
        })
    
    df = pd.DataFrame(data)
    
    # 메모리에 엑셀 파일 생성
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='메시지목록')
        
        # 워크시트 스타일링
        workbook = writer.book
        worksheet = writer.sheets['메시지목록']
        
        # 헤더 스타일
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
        
        # 열 너비 조정
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    return output.getvalue()

def create_csv_download_content(generated_messages):
    """CSV 다운로드 컨텐츠 생성"""
    data = []
    for group_id, message_data in generated_messages.items():
        group_info = message_data['group_info']
        
        data.append({
            '그룹ID': group_id,
            '팀명': group_info['team_name'],
            '발송그룹': group_info['sender_group'],
            '발송인': group_info['sender'],
            '연락처': group_info.get('contact', ''),
            '그룹멤버': ', '.join(group_info['members']),
            '인원수': group_info['group_size'],
            '총잔금': group_info.get('total_balance', '')
        })
    
    df = pd.DataFrame(data)
    return df.to_csv(index=False, encoding='utf-8-sig')

def show_error_details(error, context=""):
    """상세 오류 정보 표시"""
    st.error(f"❌ **오류 발생** {context}")
    
    with st.expander("🔧 상세 오류 정보", expanded=False):
        st.code(f"""
오류 타입: {type(error).__name__}
오류 메시지: {str(error)}
        """)
        
        # 일반적인 해결 방법 제시
        if "KeyError" in str(type(error)):
            st.markdown("""
            **💡 해결 방법:**
            - 템플릿의 변수명이 매핑된 변수와 일치하는지 확인
            - 매핑 설정에서 누락된 컬럼이 있는지 확인
            """)
        elif "ValueError" in str(type(error)):
            st.markdown("""
            **💡 해결 방법:**
            - 숫자 포맷팅({변수:,})을 문자 변수에 사용했는지 확인
            - 엑셀 데이터의 형식이 올바른지 확인
            """)
        elif "FileNotFoundError" in str(type(error)):
            st.markdown("""
            **💡 해결 방법:**
            - 파일이 제대로 업로드되었는지 확인
            - 파일 형식이 .xlsx 또는 .xls인지 확인
            """)

def create_help_sidebar():
    """도움말 사이드바 생성"""
    with st.sidebar:
        st.markdown("---")
        st.markdown("## 💡 도움말")
        
        with st.expander("📁 파일 업로드", expanded=False):
            st.markdown("""
            - 지원 형식: .xlsx, .xls
            - 최대 크기: 50MB
            - 고정 정보와 테이블이 같은 시트에 있어야 함
            """)
        
        with st.expander("🔧 매핑 설정", expanded=False):
            st.markdown("""
            - 셀 주소: A1, B2, C3 형태로 입력
            - 필수 컬럼: 팀명, 발송그룹, 이름
            - 헤더 행 번호는 1부터 시작
            """)
        
        with st.expander("📝 템플릿 작성", expanded=False):
            st.markdown("""
            - 변수 사용: {변수명}
            - 숫자 포맷: {금액:,}원
            - 미리보기로 확인 가능
            """)
        
        with st.expander("🚨 문제 해결", expanded=False):
            st.markdown("""
            - 변수 오류: 매핑 설정 확인
            - 포맷 오류: 숫자/문자 변수 구분
            - 파일 오류: 엑셀 파일 형식 확인
            """)

def show_progress_indicator(current_step, total_steps, step_names):
    """진행 상황 표시기"""
    progress = current_step / total_steps
    st.progress(progress)
    
    col1, col2 = st.columns([3, 1])
    with col1:
        if current_step <= len(step_names):
            st.write(f"**현재 단계:** {step_names[current_step-1]}")
    with col2:
        st.write(f"{current_step}/{total_steps}")

def format_currency(amount):
    """통화 포맷팅"""
    try:
        return f"{int(amount):,}원"
    except (ValueError, TypeError):
        return "0원"
    
import re

def generate_variable_name(header):
    """엑셀 헤더를 기반으로 유효한 파이썬 변수명을 생성합니다."""
    # 한글, 영문, 숫자만 남기고 나머지는 공백으로 변경
    korean = re.sub(r'[^가-힣]', ' ', header).strip()
    english = re.sub(r'[^a-zA-Z]', ' ', header).strip()

    # 영문 우선, 없으면 한글 사용
    if english:
        # 영문을 소문자로 바꾸고 공백을 '_'로 변경
        var_name = re.sub(r'\s+', '_', english.lower())
    elif korean:
        # 간단한 영어로 변환 (예시)
        translation_map = {
            '번호': 'num', '팀': 'team', '인원': 'count', '이름': 'name', '연락처': 'contact',
            '그룹': 'group', '상품가': 'price', '항공권': 'ticket', '제외': 'exclude',
            '금액': 'amount', '두바이': 'dubai', '체류': 'stay', '관광비': 'tour_fee',
            '코카서스': 'caucasus', '취소': 'cancel', '수수료': 'fee', '환불금': 'refund',
            '안내': 'info'
        }
        words = korean.split()
        translated_words = [translation_map.get(word, word) for word in words]
        var_name = '_'.join(translated_words)
    else:
        var_name = "variable"

    # 최종 정리: 맨 앞 숫자는 'var_'로 시작, 유효하지 않은 문자 제거, 길이 제한
    var_name = re.sub(r'[^a-zA-Z0-9_]', '', var_name)
    if var_name and var_name[0].isdigit():
        var_name = 'var_' + var_name
    return var_name[:50] if var_name else "unnamed_variable"

def show_smart_template_preview(template, preview_data, excel_columns):
    """
    스마트 템플릿 미리보기 (최종 생성 로직과 동기화된 버전)
    """
    st.markdown("##### 🔍 실시간 미리보기")

    if not preview_data:
        st.warning("미리보기를 생성할 데이터가 없습니다.")
        st.text_area("Preview Area", "미리보기 생성 불가", height=450, disabled=True, label_visibility="collapsed")
        return

    try:
        # 1. 미리보기에 필요한 모든 변수를 하나의 딕셔너리로 통합
        variables = {}
        variables.update(preview_data) # preview_data에는 고정 정보와 엑셀 첫 행 데이터가 모두 포함됨

        # 2. 미리보기용 특별 계산 변수 추가 (실제 값 대신 예시 값)
        # '이름' 컬럼이 있으면 해당 이름을 사용하고, 없으면 기본값 '홍길동' 사용
        name_col_key = next((col for col in excel_columns if '이름' in col), None)
        if name_col_key:
            variables.setdefault('group_members_text', f"{variables.get(name_col_key, '홍길동')}님 외 1명")
        else:
            variables.setdefault('group_members_text', "홍길동님 외 1명")
        
        variables.setdefault('group_size', 2)
        variables.setdefault('additional_fee_per_person', 70000) # 예시 추가금

        # 3. 템플릿 태그를 치환하는 콜백 함수
        def replacer(match):
            # key는 그룹 2(컬럼명) 또는 그룹 5(변수명) 중 하나가 됨
            key = match.group(2) or match.group(5)
            formatting = match.group(3) or match.group(6)
            
            if key is None:
                return match.group(0) # 매칭되지 않은 경우 원본 반환
            
            key = key.strip()
            
            # 통합 딕셔너리에서 값 조회
            value = variables.get(key, f"❌[{key}]")
            
            if isinstance(value, str) and value.startswith("❌"):
                return value

            # 숫자 포맷팅 적용
            if formatting and ':' in formatting:
                try:
                    # 문자열 내 쉼표 등 비숫자 문자 제거 후 숫자 변환
                    num_value = float(re.sub(r'[^\d.-]', '', str(value)))
                    return f"{int(num_value):,}"
                except (ValueError, TypeError):
                    return str(value) # 변환 실패 시 원본 값 반환
            
            return str(value)

        # [컬럼:키] 또는 {키} 형태의 모든 태그를 찾는 정규식
        pattern = r'\[(컬럼):([^\]:]+)(:[^\]]*)?\]|(\{)([^}]+?)(:[^}]+)?\}'
        preview_message = re.sub(pattern, replacer, template)
        
        # 5. 최종 결과 표시
        st.text_area(
            "Preview Area",
            value=preview_message,
            height=450,
            disabled=True,
            label_visibility="collapsed",
            help="엑셀의 첫 번째 데이터를 기반으로 생성된 실시간 미리보기입니다."
        )

    except Exception as e:
        st.error(f"❌ 미리보기 생성 중 오류: {str(e)}")
        import traceback
        st.code(traceback.format_exc())