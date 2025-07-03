#!/usr/bin/env python3
"""
샘플 엑셀 파일 생성 스크립트
테스트용 여행 고객 데이터를 생성합니다.
"""

import pandas as pd
import random
from datetime import datetime, timedelta

def create_sample_excel():
    """샘플 엑셀 파일 생성"""
    
    # 워크북 생성
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "여행고객정보"
    
    # 고정 정보 입력
    fixed_data = [
        ["", "", "", "상품명", "하와이 힐링 7일"],
        ["", "", "", "잔금완납일", "2024-03-15"],
        ["", "", "", "", ""],
        ["", "", "", "", ""],
        ["", "", "", "", "", "기준환율", 1300],
        ["", "", "", "", "", "현재환율", 1350],
        ["", "", "", "", "", "환율차액", 50],
        ["", "", "", "", "", "당사부담금", 20],
        ["", "", "", "", "", "", ""],
    ]
    
    # 고정 정보 스타일링
    for row_idx, row_data in enumerate(fixed_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in [4, 6] and value:  # 라벨 셀
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    # 테이블 헤더 (9행)
    headers = ["팀", "문자 발송 그룹", "이름", "연락처", "상품가", "환차금", "예약금", "항공료", "추가 금액", "1인당 잔금", "잔금 안내 금액", "입금 계좌"]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # 샘플 고객 데이터 생성
    teams = ["1팀", "2팀", "3팀", "4팀"]
    sender_groups = ["A그룹", "B그룹", "C그룹"]
    names = ["김철수", "이영희", "박민수", "정수진", "최동욱", "한소영", "임재현", "송미란", "황준호", "오세영"]
    
    # 팀별 데이터 생성
    row = 10
    
    for team in teams:
        # 각 팀마다 2-3개의 발송 그룹
        team_groups = random.sample(sender_groups, random.randint(2, 3))
        
        for group in team_groups:
            # 각 그룹마다 2-4명의 멤버
            group_size = random.randint(2, 4)
            group_names = random.sample(names, group_size)
            
            # 기본 금액 설정 (그룹 공통)
            base_price = random.choice([3200000, 3500000, 3800000])
            exchange_fee = 50000
            deposit = random.choice([500000, 800000, 1000000])
            flight_fee = random.choice([800000, 900000, 1000000])
            additional_amount = random.choice([0, 0, 0, 100000, 150000])  # 대부분 0
            
            individual_balance = base_price + exchange_fee - deposit - flight_fee + additional_amount
            total_balance = individual_balance * group_size
            
            bank_account = "우리은행 123-456-789012 (주)여행처럼"
            
            # 각 멤버 데이터 입력
            for i, name in enumerate(group_names):
                phone = f"010-{random.randint(1000, 9999)}-{random.randint(1000, 9999)}"
                
                # 첫 번째 멤버만 연락처 정보 입력 (대표자)
                if i == 0:
                    contact = phone
                else:
                    contact = ""
                
                data = [
                    team,                    # 팀
                    group,                   # 문자 발송 그룹
                    name,                    # 이름
                    contact,                 # 연락처 (대표자만)
                    base_price,              # 상품가
                    exchange_fee,            # 환차금
                    deposit,                 # 예약금
                    flight_fee,              # 항공료
                    additional_amount,       # 추가 금액
                    individual_balance,      # 1인당 잔금
                    total_balance,           # 잔금 안내 금액
                    bank_account if i == 0 else ""  # 계좌 (대표자만)
                ]
                
                for col_idx, value in enumerate(data, 1):
                    ws.cell(row=row, column=col_idx, value=value)
                
                row += 1
    
    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 파일 저장
    filename = f"sample_travel_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    
    return filename

def create_simple_sample():
    """간단한 샘플 데이터 생성 (pandas 사용)"""
    
    # 고정 정보를 별도 딕셔너리로 저장
    fixed_info = {
        "D2": "제주도 휴양 4일",
        "D3": "2024-12-20", 
        "F2": 1280,
        "F3": 1320,
        "F4": 40,
        "F5": 15
    }
    
    # 테이블 데이터 생성
    data = [
        ["1팀", "A그룹", "김민준", "010-1234-5678", 2800000, 40000, 500000, 600000, 0, 1740000, 3480000, "국민은행 123-456-789 (주)여행사"],
        ["1팀", "A그룹", "이서현", "", 2800000, 40000, 500000, 600000, 0, 1740000, 3480000, ""],
        ["1팀", "B그룹", "박준영", "010-2345-6789", 2800000, 40000, 500000, 600000, 100000, 1840000, 5520000, "국민은행 123-456-789 (주)여행사"],
        ["1팀", "B그룹", "최수빈", "", 2800000, 40000, 500000, 600000, 100000, 1840000, 5520000, ""],
        ["1팀", "B그룹", "정하윤", "", 2800000, 40000, 500000, 600000, 100000, 1840000, 5520000, ""],
        ["2팀", "A그룹", "강도현", "010-3456-7890", 2800000, 40000, 500000, 600000, 0, 1740000, 5220000, "국민은행 123-456-789 (주)여행사"],
        ["2팀", "A그룹", "윤시우", "", 2800000, 40000, 500000, 600000, 0, 1740000, 5220000, ""],
        ["2팀", "A그룹", "임채원", "", 2800000, 40000, 500000, 600000, 0, 1740000, 5220000, ""],
        ["3팀", "C그룹", "홍준서", "010-4567-8901", 2800000, 40000, 500000, 600000, 50000, 1790000, 3580000, "국민은행 123-456-789 (주)여행사"],
        ["3팀", "C그룹", "김나연", "", 2800000, 40000, 500000, 600000, 50000, 1790000, 3580000, ""],
    ]
    
    # 컬럼명
    columns = ["팀", "문자 발송 그룹", "이름", "연락처", "상품가", "환차금", "예약금", "항공료", "추가 금액", "1인당 잔금", "잔금 안내 금액", "입금 계좌"]
    
    # DataFrame 생성
    df = pd.DataFrame(data, columns=columns)
    
    # 엑셀 파일로 저장 (고정 정보 포함)
    filename = f"simple_sample_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # 빈 시트 생성하고 수동으로 데이터 입력
        workbook = writer.book
        worksheet = workbook.active
        worksheet.title = "여행정보"
        
        # 고정 정보 입력
        worksheet['D2'] = fixed_info['D2']
        worksheet['D3'] = fixed_info['D3'] 
        worksheet['F2'] = fixed_info['F2']
        worksheet['F3'] = fixed_info['F3']
        worksheet['F4'] = fixed_info['F4']
        worksheet['F5'] = fixed_info['F5']
        
        # 라벨 추가
        worksheet['C2'] = "상품명"
        worksheet['C3'] = "잔금완납일"
        worksheet['E2'] = "기준환율"
        worksheet['E3'] = "현재환율" 
        worksheet['E4'] = "환율차액"
        worksheet['E5'] = "당사부담금"
        
        # 테이블 데이터는 9행부터 시작
        for col_num, column_name in enumerate(columns, 1):
            worksheet.cell(row=9, column=col_num, value=column_name)
        
        for row_num, row_data in enumerate(data, 10):
            for col_num, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)
    
    return filename

def main():
    """메인 함수"""
    print("🎯 샘플 엑셀 파일을 생성합니다...")
    
    choice = input("""
어떤 샘플을 생성하시겠습니까?
1. 상세한 샘플 (스타일링 포함)
2. 간단한 샘플 (테스트용)
선택 (1 또는 2): """).strip()
    
    try:
        if choice == "1":
            filename = create_sample_excel()
            print(f"✅ 상세한 샘플 파일이 생성되었습니다: {filename}")
        elif choice == "2":
            filename = create_simple_sample()
            print(f"✅ 간단한 샘플 파일이 생성되었습니다: {filename}")
        else:
            print("❌ 잘못된 선택입니다. 1 또는 2를 입력해주세요.")
            return
        
        print(f"""
📋 생성된 파일 정보:
- 파일명: {filename}
- 고정정보: D2(상품명), D3(완납일), F2~F5(환율정보)
- 테이블: 9행부터 시작 (헤더 포함)
- 샘플그룹: 여러 팀/발송그룹별 데이터

🚀 이 파일을 웹 애플리케이션에서 테스트해보세요!
        """)
        
    except Exception as e:
        print(f"❌ 파일 생성 중 오류 발생: {e}")

if __name__ == "__main__":
    main()